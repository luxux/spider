# -*- coding: utf-8 -*-
# @Author: koosuf
# @Date:   2017-02-25 16:36:46
# @Last Modified by:   koosuf
# @Last Modified time: 2017-02-25 20:32:56

from spider import Postdata
from openpyxl.workbook import Workbook
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter


class ExcelDeal(object):
    title_hd = ["流域", "行政区", "河名", "站名", "时间", "水位(m)", "流量(m²/s)", "警戒水位(m)"]
    title_sk = ["流域", "行政区", "河名", "库名",
                "库水位(m)", "蓄水量(10⁹m²)", "入库(m³/s)", "堤顶高度(m)"]

    def __init__(self, ws_naem, excel_name):
        self.wb = Workbook()
        self.ws_naem = self.wb.active
        self.excel_name = excel_name

    def deal_data(self, tree):
        output_list = []
        tr_list = tree[0].xpath('//tr')
        if tree[1] == "hd":
            for i in tr_list:
                Watershed = i.xpath('.//td')[0].xpath('./text()')[0]
                Admin_area = i.xpath('.//td')[1].xpath('./text()')[0]
                River_name = i.xpath('.//td')[2].xpath('./text()')[0]
                Site_name = i.xpath('.//td')[3].xpath('./font/text()')[0]
                Time = i.xpath('.//td')[4].xpath('./text()')[0]
                Water_level = i.xpath('.//td')[5].xpath('./font/text()')[0]
                Water_level_flag = i.xpath(
                    './/td')[5].xpath('./font/text()')[1]
                Water_flow = i.xpath('.//td')[6].xpath('./text()')[0]
                Alert_water = i.xpath('.//td')[7].xpath('./text()')[0]

                Watershed = Watershed.encode(
                    'utf-8').decode('unicode_escape'),
                Admin_area = Admin_area.encode(
                    'utf-8').decode('unicode_escape'),
                River_name = River_name.encode(
                    'utf-8').decode('unicode_escape'),
                Site_name = Site_name.encode('utf-8').decode('unicode_escape'),
                Time = Time.encode('utf-8').decode('unicode_escape'),
                Water_level = Water_level.encode(
                    'utf-8').decode('unicode_escape'),
                Water_level_flag = Water_level_flag.encode(
                    'utf-8').decode('unicode_escape'),
                Water_flow = Water_flow.encode(
                    'utf-8').decode('unicode_escape'),
                Alert_water = Alert_water,

                output_list.append([Watershed[0], Admin_area[0], River_name[0], Site_name[0],
                                    Time[0], Water_level[0] + "   " + Water_level_flag[0], Water_flow[0], Alert_water[0]])

        elif tree[1] == "sk":
            pass
        else:
            print("error!")
        return output_list

    def deal_form(self):
        self.ws_naem .column_dimensions['C'].width = 13
        self.ws_naem .column_dimensions['D'].width = 20
        self.ws_naem .column_dimensions['E'].width = 20
        self.ws_naem .column_dimensions['F'].width = 20
        self.ws_naem .column_dimensions['G'].width = 20
        self.ws_naem .column_dimensions['H'].width = 20

    def load_excel(self, table_name, data_list):
        self.deal_form()
        self.ws_naem .freeze_panes = 'A2'
        self.ws_naem .title = table_name
        self.ws_naem .append(self.title_hd)
        for i in range(len(data_list)):
            self.ws_naem .append(data_list[i])

    def save_excel(self):
        try:
            self.wb.save(filename=self.excel_name)
        except IOError as e:
            print(e)
        pass

if __name__ == '__main__':
    jj = ExcelDeal('ws0', 'waterdata.xlsx')
    gg = Postdata()
    data = jj.deal_data(Postdata.fetch_html_data(gg, "hd"))
    jj.load_excel('2017-02-25', data)
    jj.save_excel()
